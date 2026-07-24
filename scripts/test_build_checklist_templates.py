"""Regression and guard tests for ``scripts/build_checklist_templates.py``.

Covers the issue #255 / PR #356 dashboard-degradation regression: the
governance-maturity dashboard must never silently flatten its 192 curated tier
cells into the generic ``All <tier> requirements met (see control doc).``
boilerplate when SPA data (``docs/javascripts/assessment-data.json``) is
absent, unparseable, or incomplete.

Tier columns of ``governance-maturity-dashboard.xlsx``:
    E = Baseline yes-bar, F = Recommended yes-bar, G = Regulated yes-bar
"""
from __future__ import annotations

import io
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "scripts"))

import build_checklist_templates as bct  # noqa: E402
import extract_assessment_data as ead  # noqa: E402

DASHBOARD_NAME = "governance-maturity-dashboard.xlsx"
DASHBOARDS = [
    REPO_ROOT / "assessment" / "templates" / DASHBOARD_NAME,
    REPO_ROOT / "docs" / "assessment" / "templates" / DASHBOARD_NAME,
]
TIER_COLUMN = {"baseline": 5, "recommended": 6, "regulated": 7}
STALE_311_PINPOINT = "(f)(2)(ii)(A)"
CORRECTED_311_PINPOINT = "17a-4(f)(2)"


# ── helpers ───────────────────────────────────────────────────────────────

def _manifest() -> list[dict]:
    return bct.load_manifest()


def _real_spa() -> dict[str, dict]:
    """SPA dict rebuilt from the *current tree* via the canonical extractor."""
    data = ead.build_output()
    assert data is not None, "extract_assessment_data.build_output() failed"
    return {c["id"]: c for c in data["controls"]}


def _tier_cells(source) -> dict[str, dict[str, str]]:
    """Return ``{control_id: {tier: text}}`` for a dashboard workbook.

    ``source`` may be a filesystem path or raw ``.xlsx`` bytes (e.g. a git blob).
    """
    if isinstance(source, (bytes, bytearray)):
        wb = load_workbook(io.BytesIO(source), data_only=False)
    else:
        wb = load_workbook(source, data_only=False)
    ws = wb.active
    header_row = None
    for r in range(1, 20):
        if str(ws.cell(row=r, column=1).value or "").strip() == "Control ID":
            header_row = r
            break
    assert header_row is not None, "'Control ID' header not found"
    out: dict[str, dict[str, str]] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        cid = ws.cell(row=r, column=1).value
        if cid is None or not str(cid).strip():
            continue
        cid = str(cid).strip()
        out[cid] = {lvl: str(ws.cell(row=r, column=col).value or "")
                    for lvl, col in TIER_COLUMN.items()}
    wb.close()
    return out


def _boilerplate_hits(cells: dict[str, dict[str, str]]) -> list[tuple[str, str]]:
    return [(cid, lvl)
            for cid, tiers in cells.items()
            for lvl, txt in tiers.items()
            if txt.strip() == bct.dashboard_boilerplate(lvl)]


def _git_blob(ref_path: str) -> bytes | None:
    res = subprocess.run(["git", "show", ref_path], cwd=str(REPO_ROOT),
                         capture_output=True)
    return res.stdout if res.returncode == 0 else None


# ── committed-artifact invariants ─────────────────────────────────────────

def test_committed_dashboards_have_zero_boilerplate() -> None:
    manifest_count = len(_manifest())
    for path in DASHBOARDS:
        cells = _tier_cells(path)
        assert len(cells) == manifest_count, (
            f"{path.name}: {len(cells)} control rows, expected {manifest_count}")
        hits = _boilerplate_hits(cells)
        assert not hits, f"{path.name}: generic boilerplate tier cells: {hits}"


def test_all_controls_retain_curated_tier_content() -> None:
    for path in DASHBOARDS:
        cells = _tier_cells(path)
        for cid, tiers in cells.items():
            for lvl, txt in tiers.items():
                assert txt.strip(), f"{path.name}: {cid}[{lvl}] is empty"
                assert txt.strip() != bct.dashboard_boilerplate(lvl), (
                    f"{path.name}: {cid}[{lvl}] is generic boilerplate")


def test_dashboard_control_311_has_corrected_citation() -> None:
    for path in DASHBOARDS:
        cells = _tier_cells(path)
        assert "3.11" in cells, f"{path.name}: control 3.11 missing"
        blob = " ".join(cells["3.11"].values())
        assert CORRECTED_311_PINPOINT in blob, (
            f"{path.name}: 3.11 missing corrected citation: {blob!r}")
        assert STALE_311_PINPOINT not in blob, (
            f"{path.name}: 3.11 still carries stale {STALE_311_PINPOINT}: {blob!r}")


def test_both_mirrored_dashboards_equivalent() -> None:
    assessment_cells, docs_cells = (_tier_cells(p) for p in DASHBOARDS)
    assert assessment_cells == docs_cells, "mirrored dashboards diverge"


# ── freshness: committed dashboards == fresh regeneration from current SPA ──

def test_committed_dashboards_match_fresh_regeneration(tmp_path) -> None:
    spa = _real_spa()
    fresh_path = tmp_path / DASHBOARD_NAME
    bct.build_dashboard(fresh_path, _manifest(), spa)
    fresh_cells = _tier_cells(fresh_path)
    for path in DASHBOARDS:
        assert _tier_cells(path) == fresh_cells, (
            f"{path.name} is stale or degraded versus a fresh SPA "
            "regeneration. Run: python scripts/extract_assessment_data.py && "
            "python scripts/build_checklist_templates.py")


# ── versus current main: zero curated -> boilerplate degradation ───────────

def test_no_curated_to_boilerplate_regression_versus_main() -> None:
    blob = _git_blob("origin/main:assessment/templates/" + DASHBOARD_NAME)
    if blob is None:
        pytest.skip("origin/main dashboard blob unavailable in this checkout")
    main_cells = _tier_cells(blob)
    live_cells = _tier_cells(DASHBOARDS[0])
    degraded: list[str] = []
    for cid, tiers in main_cells.items():
        for lvl, main_txt in tiers.items():
            if main_txt.strip() == bct.dashboard_boilerplate(lvl):
                continue  # main itself had no curated content here
            live_txt = live_cells.get(cid, {}).get(lvl, "")
            if live_txt.strip() == bct.dashboard_boilerplate(lvl):
                degraded.append(f"{cid}[{lvl}]")
    assert not degraded, (
        f"{len(degraded)} tier cell(s) degraded curated content -> "
        f"boilerplate versus main: {degraded}")


# ── root-cause guard: absent / malformed / incomplete SPA fails loudly ─────

def test_guard_raises_when_spa_absent(tmp_path) -> None:
    missing = tmp_path / "does-not-exist.json"
    with pytest.raises(SystemExit) as exc:
        bct.require_complete_dashboard_spa(_manifest(), {}, spa_path=missing)
    assert "absent" in str(exc.value).lower()


def test_load_spa_returns_empty_on_malformed_then_guard_raises(
        tmp_path, monkeypatch) -> None:
    bad = tmp_path / "assessment-data.json"
    bad.write_text("{ this is : not valid json", encoding="utf-8")
    monkeypatch.setattr(bct, "SPA_DATA", bad)
    assert bct.load_spa() == {}, "malformed SPA should parse to {}"
    with pytest.raises(SystemExit) as exc:
        bct.require_complete_dashboard_spa(_manifest(), bct.load_spa(), spa_path=bad)
    msg = str(exc.value).lower()
    assert "unparseable" in msg or "empty" in msg


def test_guard_raises_when_spa_incomplete(tmp_path) -> None:
    spa = _real_spa()
    spa.pop("3.11", None)  # drop one control -> its 3 tier cells would degrade
    existing = tmp_path / "assessment-data.json"
    existing.write_text("{}", encoding="utf-8")  # exists: skip the absence check
    with pytest.raises(SystemExit) as exc:
        bct.require_complete_dashboard_spa(_manifest(), spa, spa_path=existing)
    text = str(exc.value)
    assert "3.11" in text and "incomplete" in text.lower()


def test_guard_passes_on_current_tree(tmp_path) -> None:
    existing = tmp_path / "assessment-data.json"
    existing.write_text("{}", encoding="utf-8")
    # Complete, current SPA -> no SystemExit.
    bct.require_complete_dashboard_spa(_manifest(), _real_spa(), spa_path=existing)


def test_empty_spa_flags_every_tier_cell() -> None:
    """With no SPA every tier cell is boilerplate; the detector must catch the
    full-flatten case (192 cells for 64 controls)."""
    offenders = bct.dashboard_boilerplate_cells(_manifest(), {})
    assert len(offenders) == len(_manifest()) * 3


def test_build_dashboard_regeneration_is_boilerplate_free(tmp_path) -> None:
    """End-to-end: building from complete current SPA yields zero boilerplate."""
    out = tmp_path / DASHBOARD_NAME
    bct.build_dashboard(out, _manifest(), _real_spa())
    assert not _boilerplate_hits(_tier_cells(out))


# ── Thread PRRT_kwDORX7m3c6TagWh: standalone generation on a clean checkout ──
# The documented `python scripts/build_checklist_templates.py` command must work
# without a prior extractor run, because docs/javascripts/assessment-data.json is
# gitignored and absent on a clean checkout — while keeping the no-degradation
# guarantee (invalid canonical data still fails before writing).


def test_load_spa_from_extractor_is_complete() -> None:
    """The in-memory canonical rebuild yields complete, boilerplate-free tier
    data, so the standalone command works without the gitignored SPA file."""
    spa = bct.load_spa_from_extractor()
    assert spa, "extractor produced no SPA controls"
    offenders = bct.dashboard_boilerplate_cells(_manifest(), spa)
    assert not offenders, f"in-memory SPA would degrade tier cells: {offenders}"


def test_resolve_dashboard_spa_rebuilds_when_file_absent(
        tmp_path, monkeypatch) -> None:
    """Clean checkout: with the SPA file absent, resolve rebuilds in-memory and
    passes the no-degradation guard (no SystemExit), sourced from the extractor."""
    monkeypatch.setattr(bct, "SPA_DATA", tmp_path / "assessment-data.json")
    assert not bct.SPA_DATA.exists()
    spa, source = bct.resolve_dashboard_spa(_manifest())
    assert spa and "in-memory" in source
    assert not bct.dashboard_boilerplate_cells(_manifest(), spa)


def test_resolve_dashboard_spa_fails_before_write_on_invalid_canonical(
        tmp_path, monkeypatch) -> None:
    """Clean checkout with an unusable canonical source (extractor yields no
    controls, e.g. a control doc fails to parse) must still fail *before* writing
    rather than degrade the dashboard to boilerplate."""
    monkeypatch.setattr(bct, "SPA_DATA", tmp_path / "assessment-data.json")
    monkeypatch.setattr(bct, "load_spa_from_extractor", lambda: {})
    with pytest.raises(SystemExit) as exc:
        bct.resolve_dashboard_spa(_manifest())
    assert "empty or unparseable" in str(exc.value).lower()


def test_main_builds_complete_dashboard_on_clean_checkout(
        tmp_path, monkeypatch) -> None:
    """End-to-end reviewer scenario: on a clean checkout (no assessment-data.json)
    ``python scripts/build_checklist_templates.py`` still writes a complete,
    boilerplate-free 192-cell dashboard (and valid role checklists) via the
    in-memory canonical rebuild — no hidden extractor step required."""
    out = tmp_path / "templates"
    docs_mirror = tmp_path / "docs-templates"
    monkeypatch.setattr(bct, "SPA_DATA", tmp_path / "assessment-data.json")
    monkeypatch.setattr(bct, "OUT_DIR", out)
    monkeypatch.setattr(bct, "DOCS_MIRROR", docs_mirror)
    assert not bct.SPA_DATA.exists()

    assert bct.main() == 0
    dash = out / DASHBOARD_NAME
    assert dash.exists(), "dashboard not written on clean checkout"
    cells = _tier_cells(dash)
    assert len(cells) == len(_manifest())
    assert not _boilerplate_hits(cells), "clean-checkout dashboard has boilerplate"
    # Role checklists remain valid and the docs mirror is written too.
    xlsx = list(out.glob("*.xlsx"))
    assert len(xlsx) == len(bct.ROLE_FILES) + 1  # roles + dashboard
    assert (docs_mirror / DASHBOARD_NAME).exists()
