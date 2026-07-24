"""Generate FSI-CopilotGov Excel checklist templates and governance dashboard.

Reads ``assessment/manifest/controls.json`` for control content and sources the
governance-maturity dashboard's tier columns from
``docs/javascripts/assessment-data.json``. That SPA data file is gitignored and
absent on a clean checkout, so when it is missing this script rebuilds the tier
data in-memory via ``extract_assessment_data.build_output`` (the same production
logic that writes the published file). The documented standalone command below
therefore works on a clean checkout without a separate extractor step, while the
no-degradation guard still fails loudly before writing if the canonical source
is absent/empty/incomplete. Emits role-specific ``.xlsx`` checklists plus a
governance maturity dashboard into ``assessment/templates/``.

The MkDocs hook ``scripts/hooks/copy_assessment_data.py`` publishes the
files to ``site/assessment/templates/`` so the download links in
``docs/getting-started/checklist.md`` resolve on the built site.

Usage::

    python scripts/build_checklist_templates.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO_ROOT = Path(__file__).resolve().parents[1]
MANIFEST = REPO_ROOT / "assessment" / "manifest" / "controls.json"
SPA_DATA = REPO_ROOT / "docs" / "javascripts" / "assessment-data.json"
OUT_DIR = REPO_ROOT / "assessment" / "templates"
# Mirror into docs/ so mkdocs link resolution (and local serve) pick up
# the real XLSX files; the mkdocs post_build hook also copies from
# OUT_DIR into site/assessment/templates/ for the published artifact.
DOCS_MIRROR = REPO_ROOT / "docs" / "assessment" / "templates"

FRAMEWORK_VERSION = "FSI Copilot Governance Framework v1.8.0"

# ── Role → checklist file mapping ─────────────────────────────────────────────
# Each entry: (output filename, display role label, manifest role key or None).
# Roles not represented in assessment/manifest/controls.json are driven by the
# explicit ROLE_CONTROL_OVERRIDES map below (source: role homework pages +
# extract_assessment_data.py ROLE_CONTROLS alignment for supervision/records).
ROLE_FILES = [
    ("ai-administrator-checklist.xlsx",
     "AI Administrator", "AI Governance Lead"),
    ("m365-global-admin-or-copilot-admin-checklist.xlsx",
     "M365 Global Admin / Copilot Admin", "M365 Global Admin"),
    ("entra-global-admin-checklist.xlsx",
     "Entra Global Admin", "Entra Global Admin"),
    ("sharepoint-admin-checklist.xlsx",
     "SharePoint Admin", "SharePoint Admin"),
    ("purview-compliance-admin-checklist.xlsx",
     "Purview Compliance Admin", "Purview Compliance Admin"),
    ("teams-admin-checklist.xlsx",
     "Teams Admin", "Teams Admin"),
    ("security-admin-checklist.xlsx",
     "Security Admin", "Security Admin"),
    ("compliance-officer-checklist.xlsx",
     "Compliance Officer", None),
    ("exchange-online-admin-checklist.xlsx",
     "Exchange Online Admin", "Exchange Online Admin"),
    ("internal-audit-checklist.xlsx",
     "Internal Audit", None),
    ("privacy-officer-checklist.xlsx",
     "Privacy Officer", None),
    ("records-manager-checklist.xlsx",
     "Records Manager", None),
    ("vendor-third-party-risk-manager-checklist.xlsx",
     "Vendor / Third-Party Risk Manager", None),
    ("governance-lead-checklist.xlsx",
     "Governance Lead", None),
]

# Role-specific control overlays for templates where manifest role metadata is
# intentionally incomplete (e.g., TODO role assignment placeholders) or absent.
ROLE_CONTROL_OVERRIDES = {
    # Maintain direct supervision/records controls in compliance views.
    "compliance-officer-checklist.xlsx": [
       "1.10", "1.12", "2.1", "2.2", "2.10",
       "3.1", "3.2", "3.3", "3.4", "3.5", "3.6", "3.7",
       "3.8", "3.8a", "3.9", "3.10", "3.11", "3.12", "3.13", "3.14",
       "4.15", "4.16",
    ],
    # TODO-role controls in manifest (3.8a, 3.14) still belong in Purview ops.
    "purview-compliance-admin-checklist.xlsx": ["3.8a", "3.14"],
    # Homework-referenced roles published under docs/assessment/templates.
    "internal-audit-checklist.xlsx": ["3.1", "3.12", "3.13"],
    "privacy-officer-checklist.xlsx": ["2.5", "3.10"],
    "records-manager-checklist.xlsx": [
       "3.1", "3.2", "3.3", "3.11", "3.12", "3.14", "4.15", "4.16"
    ],
    "vendor-third-party-risk-manager-checklist.xlsx": ["1.10", "1.13"],
    "governance-lead-checklist.xlsx": ["1.1", "1.11", "1.12", "3.7", "4.9", "4.12"],
    "exchange-online-admin-checklist.xlsx": ["3.2", "3.11"],
}

# ── Styling ───────────────────────────────────────────────────────────────────
PRIMARY_DARK = "0078D4"   # FSI Microsoft Blue
PRIMARY_LIGHT = "D6E8F7"
WHITE = "FFFFFF"
LIGHT_GREY = "F4F6F8"
MED_GREY = "6B7280"

TITLE_FONT = Font(name="Aptos", bold=True, size=16, color=PRIMARY_DARK)
SUBTITLE_FONT = Font(name="Aptos", size=10, color=MED_GREY)
HEADER_FONT = Font(name="Aptos", bold=True, size=11, color=WHITE)
HEADER_FILL = PatternFill(start_color=PRIMARY_DARK,
                          end_color=PRIMARY_DARK, fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color=LIGHT_GREY,
                           end_color=LIGHT_GREY, fill_type="solid")
DATA_FONT = Font(name="Aptos", size=11)
ID_FONT = Font(name="Aptos", bold=True, size=11, color=PRIMARY_DARK)
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# ── Column schemas ────────────────────────────────────────────────────────────
CHECKLIST_HEADERS = [
    "Control ID", "Title", "Pillar", "Tier", "Surface(s)",
    "Question", "Evidence Expected", "Yes/Partial/No", "Notes",
]
CHECKLIST_WIDTHS = [12, 48, 22, 32, 26, 60, 60, 18, 40]

DASHBOARD_HEADERS = [
    "Control ID", "Title", "Pillar", "Surface",
    "Baseline yes-bar", "Recommended yes-bar", "Regulated yes-bar",
    "Status", "Score",
]
DASHBOARD_WIDTHS = [12, 48, 22, 26, 50, 50, 50, 16, 10]

STATUS_OPTIONS = '"Yes,Partial,No"'


def load_manifest() -> list[dict]:
    with MANIFEST.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def load_spa() -> dict[str, dict]:
    """Return {control_id: control_dict} from the SPA data file, or {}."""
    if not SPA_DATA.exists():
        return {}
    try:
        with SPA_DATA.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (json.JSONDecodeError, OSError):
        return {}
    return {c["id"]: c for c in data.get("controls", [])}


def pillar_surface(pillar_name: str) -> str:
    """Map a pillar name to a Copilot-surface description."""
    mapping = {
        "Readiness & Assessment": "Tenant readiness; SharePoint/OneDrive; Graph",
        "Security & Data Protection":
            "Copilot Chat; Pages; Agents; Purview DLP/Labels",
        "Compliance & Records":
            "Purview audit/eDiscovery/retention; FINRA/SEC records",
        "Operations & Lifecycle":
            "Copilot admin center; Teams/Viva; Sentinel",
    }
    return mapping.get(pillar_name, pillar_name)


def tier_text(spa_control: dict | None) -> str:
    if spa_control and spa_control.get("governanceLevels"):
        return str(spa_control["governanceLevels"])
    return "Baseline / Recommended / Regulated"


def evidence_text(manifest_control: dict, spa_control: dict | None) -> str:
    items = manifest_control.get("evidenceExpected") or []
    if not items and spa_control:
        items = [vc.get("text", "") for vc in
                 spa_control.get("verificationCriteria", [])
                 if vc.get("text")]
    items = [t.strip().rstrip("-").strip() for t in items if t and t.strip()]
    return "; ".join(items)


DASHBOARD_TIERS = ("baseline", "recommended", "regulated")


def dashboard_boilerplate(level: str) -> str:
    """Generic placeholder for a tier that has no SPA-derived rationale.

    Appearing in a *generated* dashboard means SPA data was absent or
    incomplete at build time; ``require_complete_dashboard_spa`` treats that as
    a hard error rather than let 192 curated tier cells silently flatten
    (regression from issue #255 / PR #356).
    """
    return f"All {level.title()} requirements met (see control doc)."


def yes_bar(spa_control: dict | None, level: str) -> str:
    """Derive a yes-bar rationale line for a tier from SPA levelRequirements."""
    if not spa_control:
        return dashboard_boilerplate(level)
    lr = spa_control.get("levelRequirements", {}).get(level) or {}
    rationale = lr.get("rationale")
    if rationale:
        return rationale.strip()
    reqs = lr.get("requirements") or []
    if reqs:
        return reqs[0].strip()
    return dashboard_boilerplate(level)


def build_role_to_ids(manifest: list[dict]) -> dict[str, list[str]]:
    mapping: dict[str, list[str]] = {}
    for c in manifest:
        for r in (c.get("roles") or []):
            mapping.setdefault(r, []).append(c["id"])
    return mapping


CONTROL_ID_RE = re.compile(r"^(?P<pillar>\d+)\.(?P<number>\d+)(?P<suffix>[a-z]?)$", re.IGNORECASE)


def control_id_key(control_id: str) -> tuple[int, int, str]:
    match = CONTROL_ID_RE.match(control_id.strip())
    if not match:
        return (999, 999, control_id)
    return (
        int(match.group("pillar")),
        int(match.group("number")),
        match.group("suffix").lower(),
    )


def select_ids(filename: str, role_key: str | None,
               role_to_ids: dict[str, list[str]], all_ids: list[str]) -> list[str]:
    all_id_set = set(all_ids)
    ids = set(role_to_ids.get(role_key, [])) if role_key else set()
    ids.update(
        cid for cid in ROLE_CONTROL_OVERRIDES.get(filename, [])
        if cid in all_id_set
    )
    return sorted_by_natural_id(list(ids))


def sorted_by_natural_id(ids: list[str]) -> list[str]:
    return sorted(ids, key=control_id_key)


# ── Worksheet builders ───────────────────────────────────────────────────────

def _write_title(ws, title: str, subtitle: str, cols: int) -> None:
    last = get_column_letter(cols)
    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6


def _write_headers(ws, headers: list[str], widths: list[int],
                   header_row: int) -> None:
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 28


def _write_data_row(ws, row: int, values: list, alt: bool) -> None:
    fill = ALT_ROW_FILL if alt else PatternFill()
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.border = THIN
        cell.fill = fill
        if col_idx == 1:
            cell.font = ID_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 42


def build_checklist(out_path: Path, role_label: str, controls: list[dict],
                    spa: dict[str, dict]) -> int:
    wb = Workbook()
    ws = wb.active
    safe_title = role_label.replace("/", "-")[:31]
    ws.title = safe_title

    subtitle = (f"Role: {role_label}  |  {FRAMEWORK_VERSION}  |  "
                f"{len(controls)} controls")
    _write_title(ws, f"{role_label} Checklist", subtitle,
                 len(CHECKLIST_HEADERS))

    header_row = 4
    _write_headers(ws, CHECKLIST_HEADERS, CHECKLIST_WIDTHS, header_row)

    data_start = header_row + 1
    row = data_start
    for i, c in enumerate(controls):
        spa_c = spa.get(c["id"])
        question = c.get("manual_question") or c.get("name") or c.get("title")
        pillar_label = f"P{c['pillar']} – {c.get('pillar_name', '')}"
        _write_data_row(ws, row, [
            c["id"],
            c.get("name") or c.get("title", ""),
            pillar_label,
            tier_text(spa_c),
            pillar_surface(c.get("pillar_name", "")),
            question,
            evidence_text(c, spa_c),
            "",  # Yes/Partial/No
            "",  # Notes
        ], alt=(i % 2 == 1))
        row += 1

    data_end = row - 1

    # Dropdown on the Yes/Partial/No column (H)
    dv = DataValidation(type="list", formula1=STATUS_OPTIONS,
                        allow_blank=True, showErrorMessage=True)
    dv.error = "Select Yes, Partial, or No"
    dv.errorTitle = "Invalid value"
    dv.prompt = "Select response"
    dv.promptTitle = "Status"
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"H{data_start}:H{data_end}")

    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(CHECKLIST_HEADERS))}{data_end}")

    wb.save(out_path)
    return len(controls)


def build_dashboard(out_path: Path, manifest: list[dict],
                    spa: dict[str, dict]) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = "Governance Dashboard"

    subtitle = (f"{FRAMEWORK_VERSION}  |  Governance maturity across "
                f"all {len(manifest)} controls")
    _write_title(ws, "Governance Maturity Dashboard", subtitle,
                 len(DASHBOARD_HEADERS))

    header_row = 4
    _write_headers(ws, DASHBOARD_HEADERS, DASHBOARD_WIDTHS, header_row)

    data_start = header_row + 1
    row = data_start
    controls = sorted(manifest, key=lambda c: control_id_key(c["id"]))
    for i, c in enumerate(controls):
        spa_c = spa.get(c["id"])
        pillar_label = f"P{c['pillar']} – {c.get('pillar_name', '')}"
        _write_data_row(ws, row, [
            c["id"],
            c.get("name") or c.get("title", ""),
            pillar_label,
            pillar_surface(c.get("pillar_name", "")),
            yes_bar(spa_c, "baseline"),
            yes_bar(spa_c, "recommended"),
            yes_bar(spa_c, "regulated"),
            "",  # Status (blank)
            None,  # Score — set via formula below
        ], alt=(i % 2 == 1))
        status_cell = f"H{row}"
        ws.cell(row=row, column=9).value = (
            f'=IF({status_cell}="Yes",1,'
            f'IF({status_cell}="Partial",0.5,0))')
        ws.cell(row=row, column=9).alignment = Alignment(
            horizontal="center", vertical="center")
        row += 1

    data_end = row - 1

    dv = DataValidation(type="list", formula1=STATUS_OPTIONS,
                        allow_blank=True, showErrorMessage=True)
    dv.prompt = "Select response"
    dv.promptTitle = "Status"
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"H{data_start}:H{data_end}")

    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(DASHBOARD_HEADERS))}{data_end}")

    wb.save(out_path)
    return len(controls)


def _rel(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def dashboard_boilerplate_cells(manifest: list[dict],
                                spa: dict[str, dict]) -> list[tuple[str, str]]:
    """Return ``[(control_id, tier)]`` whose dashboard yes-bar would fall back
    to the generic boilerplate placeholder — i.e. cells that would silently
    lose curated tier content because SPA data is missing or incomplete."""
    offenders: list[tuple[str, str]] = []
    for c in manifest:
        spa_c = spa.get(c["id"])
        for level in DASHBOARD_TIERS:
            if yes_bar(spa_c, level) == dashboard_boilerplate(level):
                offenders.append((c["id"], level))
    return offenders


def _require_complete_spa_data(manifest: list[dict],
                               spa: dict[str, dict],
                               *, source: str) -> None:
    """Fail *before* writing when resolved SPA tier data is empty/unparseable or
    incomplete. Shared by the file-based guard and the in-memory (clean-checkout)
    rebuild path so neither can silently degrade the 192 curated tier cells.

    ``source`` describes where ``spa`` came from (a file path or the in-memory
    extractor) and is used only in the error message."""
    if not spa:
        raise SystemExit(
            "ERROR: cannot build governance dashboard — SPA data at "
            f"{source} is empty or unparseable.\n"
            "       Regenerate it:  python scripts/extract_assessment_data.py"
        )
    offenders = dashboard_boilerplate_cells(manifest, spa)
    if offenders:
        preview = ", ".join(f"{cid}[{tier}]" for cid, tier in offenders[:8])
        more = "" if len(offenders) <= 8 else f" (+{len(offenders) - 8} more)"
        raise SystemExit(
            "ERROR: refusing to build governance dashboard — SPA data is "
            f"incomplete for {len(offenders)} tier cell(s) that would degrade "
            f"to generic boilerplate: {preview}{more}.\n"
            "       Regenerate SPA data:  python scripts/extract_assessment_data.py"
        )


def require_complete_dashboard_spa(manifest: list[dict],
                                   spa: dict[str, dict],
                                   spa_path: Path = SPA_DATA) -> None:
    """File-based dashboard guard: fail loudly *before* generating the dashboard
    when the SPA data file is absent, then delegate the empty/incomplete checks
    to :func:`_require_complete_spa_data`.

    The governance-maturity dashboard's three tier columns are sourced solely
    from ``docs/javascripts/assessment-data.json``. That file is untracked and
    absent in a clean checkout; :func:`resolve_dashboard_spa` rebuilds it
    in-memory in that case, so this hard file guard remains for the file path and
    for callers/tests that pass an explicit ``spa_path``. Role checklists do not
    consume SPA tier data, so this guard is scoped to the dashboard only."""
    if not spa_path.exists():
        raise SystemExit(
            "ERROR: cannot build governance dashboard — SPA data is absent at "
            f"{_rel(spa_path)}.\n"
            "       Generate it first:  python scripts/extract_assessment_data.py"
        )
    _require_complete_spa_data(manifest, spa, source=_rel(spa_path))


def load_spa_from_extractor() -> dict[str, dict]:
    """Rebuild the SPA control map in-memory via the canonical extractor.

    ``docs/javascripts/assessment-data.json`` is gitignored and absent on a
    clean checkout, so this lets the documented standalone command
    ``python scripts/build_checklist_templates.py`` produce a complete dashboard
    without a separate ``extract_assessment_data.py`` invocation. It reuses the
    same production ``build_output`` logic that writes the published file, so the
    result is byte-for-byte equivalent to loading the file. Returns ``{}`` when
    the extractor cannot produce controls (its own validation failed), so the
    caller's guard fails before writing rather than degrading the dashboard."""
    scripts_dir = str(Path(__file__).resolve().parent)
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
    try:
        import extract_assessment_data as ead
    except ImportError:
        return {}
    data = ead.build_output()
    if not data:
        return {}
    return {c["id"]: c for c in data.get("controls", []) if c.get("id")}


def resolve_dashboard_spa(manifest: list[dict]) -> tuple[dict[str, dict], str]:
    """Resolve SPA tier data for the dashboard and validate it before writing.

    Prefers the published ``docs/javascripts/assessment-data.json``; on a clean
    checkout (that file is gitignored/absent) rebuilds the data in-memory via
    :func:`load_spa_from_extractor`. Either way the no-degradation guard runs, so
    an absent/empty/incomplete canonical source fails *before* any workbook is
    written. Returns ``(spa, source_label)``."""
    spa = load_spa()
    if spa or SPA_DATA.exists():
        # File present (or present-but-unusable): keep the strict file guard,
        # which surfaces empty/unparseable/incomplete published data.
        require_complete_dashboard_spa(manifest, spa)
        return spa, _rel(SPA_DATA)
    # Clean checkout: the SPA file is gitignored and absent. Rebuild the tier
    # data canonically in-memory, then validate it before writing.
    spa = load_spa_from_extractor()
    source = "extract_assessment_data.build_output() (in-memory)"
    _require_complete_spa_data(manifest, spa, source=source)
    return spa, source


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_MIRROR.mkdir(parents=True, exist_ok=True)
    manifest = load_manifest()
    # Resolve SPA tier data — the published file, or a canonical in-memory rebuild
    # on a clean checkout where the file is gitignored/absent — and fail loudly
    # before writing if it is absent/empty/incomplete. Never silently regenerate
    # 192 boilerplate cells (issue #255 / PR #356 regression).
    spa, spa_source = resolve_dashboard_spa(manifest)
    by_id = {c["id"]: c for c in manifest}
    all_ids = [c["id"] for c in manifest]
    role_to_ids = build_role_to_ids(manifest)

    print(f"Loaded {len(manifest)} controls from {MANIFEST.relative_to(REPO_ROOT)}")
    print(f"Enriching with SPA data ({len(spa)} controls, source: {spa_source})")

    import shutil

    def _mirror(path: Path) -> None:
        shutil.copyfile(path, DOCS_MIRROR / path.name)

    for filename, role_label, manifest_key in ROLE_FILES:
        ids = select_ids(filename, manifest_key, role_to_ids, all_ids)
        controls = [by_id[cid] for cid in ids if cid in by_id]
        out = OUT_DIR / filename
        count = build_checklist(out, role_label, controls, spa)
        _mirror(out)
        print(f"  wrote {filename:58s} ({count:2d} controls)")

    dash = OUT_DIR / "governance-maturity-dashboard.xlsx"
    count = build_dashboard(dash, manifest, spa)
    _mirror(dash)
    print(f"  wrote {dash.name:58s} ({count:2d} controls)")

    return 0


if __name__ == "__main__":
    sys.exit(main())
